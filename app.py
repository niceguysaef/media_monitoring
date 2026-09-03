import hmac
import json
import os
import random
import re
import sqlite3
import threading
import urllib.error
import urllib.parse
import urllib.request
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date, timedelta
from hashlib import sha256
from io import BytesIO
from pathlib import Path
from time import time
from typing import Any, Callable, Dict, List, Optional, Tuple

from dotenv import load_dotenv
from fastapi import FastAPI, HTTPException, Query, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, Response
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field
from pdf_report import build_pdf_report
from pptx_report import build_pptx_report


load_dotenv()

BRAND24_BASE_URL = "https://api-data.brand24.com"
UPSTREAM_TIMEOUT_SECONDS = int(os.getenv("UPSTREAM_TIMEOUT_SECONDS", "10"))
UPSTREAM_MAX_CONCURRENCY = max(1, int(os.getenv("UPSTREAM_MAX_CONCURRENCY", "8")))
UPSTREAM_HOURLY_REQUEST_LIMIT = max(1, int(os.getenv("UPSTREAM_HOURLY_REQUEST_LIMIT", "250")))
UPSTREAM_DAILY_REQUEST_LIMIT = max(1, int(os.getenv("UPSTREAM_DAILY_REQUEST_LIMIT", "1500")))
RATE_LIMIT_DB_PATH = Path(os.getenv("RATE_LIMIT_DB_PATH", "data/request_limits.sqlite3"))
TRUST_PROXY_HEADERS = os.getenv("TRUST_PROXY_HEADERS", "false").casefold() == "true"
TRIAL_ACCESS_PASSWORD = os.getenv("TRIAL_ACCESS_PASSWORD", "")
TRIAL_SESSION_SECRET = os.getenv("TRIAL_SESSION_SECRET", "") or (
    sha256(f"trial-session:{TRIAL_ACCESS_PASSWORD}".encode("utf-8")).hexdigest()
    if TRIAL_ACCESS_PASSWORD
    else ""
)
TRIAL_SESSION_HOURS = max(1, int(os.getenv("TRIAL_SESSION_HOURS", "72")))
AUTH_COOKIE_SECURE = os.getenv("AUTH_COOKIE_SECURE", "false").casefold() == "true"
AUTH_COOKIE_SAMESITE = os.getenv("AUTH_COOKIE_SAMESITE", "lax").casefold()
if AUTH_COOKIE_SAMESITE not in {"lax", "strict", "none"}:
    AUTH_COOKIE_SAMESITE = "lax"
AUTH_COOKIE_NAME = "zestar_trial_session"
CORS_ALLOWED_ORIGINS = [
    origin.strip()
    for origin in os.getenv(
        "CORS_ALLOWED_ORIGINS",
        "http://localhost:3000,http://127.0.0.1:3000",
    ).split(",")
    if origin.strip()
]
DASHBOARD_CACHE_TTL_SECONDS = 180
PROJECT_CACHE_TTL_SECONDS = 60

RATE_LIMITS = {
    "all_requests_client": {"limit": 90, "window": 60},
    "all_requests_global": {"limit": 240, "window": 60},
    "dashboard": {"limit": 12, "window": 60},
    "mentions_read": {"limit": 30, "window": 60},
    "sources_read": {"limit": 10, "window": 60},
    "topics_read": {"limit": 10, "window": 60},
    "exports_create": {"limit": 2, "window": 600},
    "projects_read": {"limit": 12, "window": 60},
    "auth_login": {"limit": 5, "window": 900},
}

_upstream_slots = threading.BoundedSemaphore(UPSTREAM_MAX_CONCURRENCY)
_dashboard_cache: Dict[str, Tuple[float, Dict[str, Any]]] = {}
_projects_cache: Optional[Tuple[float, List[Dict[str, str]]]] = None
_categories_cache: Optional[Tuple[float, List[str]]] = None

app = FastAPI(
    title="Media Monitoring Demo API",
    description="Backend-for-frontend service for the media intelligence dashboard.",
    version="0.1.0",
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=CORS_ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["Content-Disposition"],
)


def _client_id(request: Request) -> str:
    forwarded_for = request.headers.get("x-forwarded-for")
    if TRUST_PROXY_HEADERS and forwarded_for:
        return forwarded_for.split(",")[0].strip()
    return request.client.host if request.client else "unknown"


def _trial_session_value(expires_at: int) -> str:
    signature = hmac.new(
        TRIAL_SESSION_SECRET.encode("utf-8"),
        f"trial:{expires_at}".encode("utf-8"),
        sha256,
    ).hexdigest()
    return f"{expires_at}.{signature}"


def _is_trial_authenticated(request: Request) -> bool:
    if not TRIAL_ACCESS_PASSWORD:
        return True
    raw_cookie = request.cookies.get(AUTH_COOKIE_NAME, "")
    try:
        expires_text, provided_signature = raw_cookie.split(".", 1)
        expires_at = int(expires_text)
    except (TypeError, ValueError):
        return False
    if expires_at <= int(time()):
        return False
    expected_signature = _trial_session_value(expires_at).split(".", 1)[1]
    return hmac.compare_digest(provided_signature, expected_signature)


def _rate_limit_connection() -> sqlite3.Connection:
    RATE_LIMIT_DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    connection = sqlite3.connect(RATE_LIMIT_DB_PATH, timeout=3)
    connection.execute(
        "CREATE TABLE IF NOT EXISTS rate_events ("
        "scope TEXT NOT NULL, identity TEXT NOT NULL, occurred REAL NOT NULL)"
    )
    connection.execute(
        "CREATE INDEX IF NOT EXISTS rate_events_lookup "
        "ON rate_events(scope, identity, occurred)"
    )
    return connection


def _check_rate_limit(bucket: str, client_id: str) -> None:
    rule = RATE_LIMITS[bucket]
    now = time()
    cutoff = now - rule["window"]
    try:
        with _rate_limit_connection() as connection:
            connection.execute("BEGIN IMMEDIATE")
            connection.execute(
                "DELETE FROM rate_events WHERE scope = ? AND identity = ? AND occurred < ?",
                (bucket, client_id, cutoff),
            )
            count, oldest = connection.execute(
                "SELECT COUNT(*), MIN(occurred) FROM rate_events WHERE scope = ? AND identity = ?",
                (bucket, client_id),
            ).fetchone()
            if count >= rule["limit"]:
                retry_after = max(1, int(rule["window"] - (now - oldest)) + 1)
                connection.rollback()
                raise HTTPException(
                    status_code=429,
                    detail=f"Too many requests. Try again in {retry_after} seconds.",
                    headers={"Retry-After": str(retry_after)},
                )
            connection.execute(
                "INSERT INTO rate_events(scope, identity, occurred) VALUES (?, ?, ?)",
                (bucket, client_id, now),
            )
    except HTTPException:
        raise
    except sqlite3.Error:
        raise HTTPException(status_code=503, detail="Request safety service is temporarily unavailable.")


def _consume_upstream_budget() -> None:
    now = time()
    budgets = (
        ("upstream_hour", 3600, UPSTREAM_HOURLY_REQUEST_LIMIT),
        ("upstream_day", 86400, UPSTREAM_DAILY_REQUEST_LIMIT),
    )
    try:
        with _rate_limit_connection() as connection:
            connection.execute("BEGIN IMMEDIATE")
            for scope, window, limit in budgets:
                connection.execute(
                    "DELETE FROM rate_events WHERE scope = ? AND identity = 'global' AND occurred < ?",
                    (scope, now - window),
                )
                count, oldest = connection.execute(
                    "SELECT COUNT(*), MIN(occurred) FROM rate_events WHERE scope = ? AND identity = 'global'",
                    (scope,),
                ).fetchone()
                if count >= limit:
                    retry_after = max(1, int(window - (now - oldest)) + 1)
                    connection.rollback()
                    raise HTTPException(
                        status_code=429,
                        detail="The monitoring request budget has been reached. Try again later.",
                        headers={"Retry-After": str(retry_after)},
                    )
            connection.executemany(
                "INSERT INTO rate_events(scope, identity, occurred) VALUES (?, 'global', ?)",
                [(scope, now) for scope, _, _ in budgets],
            )
    except HTTPException:
        raise
    except sqlite3.Error:
        raise HTTPException(status_code=503, detail="Request safety service is temporarily unavailable.")


@app.middleware("http")
async def limit_api_requests(request: Request, call_next: Callable[..., Any]) -> Response:
    # CORS preflight requests do not carry the trial session cookie. Let the
    # CORS middleware validate and answer them before enforcing authentication.
    if request.method == "OPTIONS":
        return await call_next(request)
    if request.url.path.startswith("/api/"):
        client_id = _client_id(request)
        try:
            _check_rate_limit("all_requests_client", client_id)
            _check_rate_limit("all_requests_global", "global")
        except HTTPException as error:
            return JSONResponse(
                status_code=error.status_code,
                content={"detail": error.detail},
                headers=error.headers,
            )
        public_paths = {"/api/auth/status", "/api/auth/login"}
        if request.url.path not in public_paths and not _is_trial_authenticated(request):
            return JSONResponse(
                status_code=401,
                content={"detail": "Trial access is required."},
                headers={"Cache-Control": "no-store"},
            )
    return await call_next(request)


def _cached_dashboard(cache_key: str, loader: Callable[[], Dict[str, Any]]) -> Dict[str, Any]:
    now = time()
    cached = _dashboard_cache.get(cache_key)
    if cached and now - cached[0] < DASHBOARD_CACHE_TTL_SECONDS:
        return {**cached[1], "cache": "hit"}

    value = loader()
    _dashboard_cache[cache_key] = (now, value)
    return {**value, "cache": "miss"}


def _date_range(days: int) -> List[str]:
    today = date.today()
    start = today - timedelta(days=days - 1)
    return [(start + timedelta(days=index)).isoformat() for index in range(days)]


def _white_label_message(value: Any) -> str:
    return re.sub(r"brand\s*24", "monitoring service", str(value), flags=re.IGNORECASE)


def _brand24_get(
    path: str,
    api_key: str,
    query: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    params = urllib.parse.urlencode(query or {})
    url = f"{BRAND24_BASE_URL}{path}"
    if params:
        url = f"{url}?{params}"

    request = urllib.request.Request(url, headers={"X-Api-Key": api_key})

    if not _upstream_slots.acquire(timeout=2):
        raise HTTPException(status_code=503, detail="The monitoring service is busy. Try again shortly.")

    try:
        _consume_upstream_budget()
        with urllib.request.urlopen(request, timeout=UPSTREAM_TIMEOUT_SECONDS) as response:
            return json.loads(response.read().decode("utf-8"))
    except urllib.error.HTTPError as error:
        message = error.read().decode("utf-8") or error.reason
        raise HTTPException(status_code=error.code, detail=_white_label_message(message))
    except urllib.error.URLError as error:
        raise HTTPException(status_code=502, detail=f"Monitoring service request failed: {error.reason}")
    finally:
        _upstream_slots.release()


class ExportCreate(BaseModel):
    dateFrom: date
    dateTo: date
    format: str = Field(default="xlsx", max_length=12)
    language: str = Field(default="en", max_length=8)
    reportTitle: Optional[str] = Field(default=None, max_length=120)
    organization: Optional[str] = Field(default=None, max_length=120)
    sections: List[str] = Field(
        default_factory=lambda: ["overview", "daily", "mentions", "sources", "topics", "authors", "links", "hashtags"],
        min_length=1,
        max_length=8,
    )


class PasswordLogin(BaseModel):
    password: str = Field(..., min_length=1, max_length=256)


def _mock_dashboard(brand: str, days: int) -> Dict[str, Any]:
    random.seed(f"{brand}-{days}")
    dates = _date_range(days)
    mentions = []
    sentiment = []
    reach = []
    total_mentions = 0
    total_reach = 0
    positive = 0
    negative = 0

    for index, day in enumerate(dates):
        baseline = 74 + (index % 6) * 8
        spike = 90 if index in {6, 18, 25} else 0
        count = baseline + spike + random.randint(-18, 22)
        pos = int(count * random.uniform(0.42, 0.62))
        neg = int(count * random.uniform(0.08, 0.2))
        daily_reach = count * random.randint(950, 1850)

        mentions.append({"date": day, "mentions": count})
        sentiment.append({"date": day, "positive": pos, "negative": neg, "neutral": count - pos - neg})
        reach.append({"date": day, "reach": daily_reach})

        total_mentions += count
        total_reach += daily_reach
        positive += pos
        negative += neg

    return {
        "mode": "mock",
        "brand": brand,
        "dateRange": {"from": dates[0], "to": dates[-1], "days": days},
        "kpis": {
            "mentions": total_mentions,
            "reach": total_reach,
            "sentimentScore": round(((positive - negative) / max(positive + negative, 1)) * 100),
            "shareOfVoice": 38,
            "engagementLift": 17,
        },
        "mentionsTrend": mentions,
        "sentimentTrend": sentiment,
        "reachTrend": reach,
        "sources": [
            {"name": "News", "value": 35},
            {"name": "X", "value": 24},
            {"name": "Blogs", "value": 18},
            {"name": "Forums", "value": 13},
            {"name": "TikTok", "value": 10},
        ],
        "topics": [
            {"topic": "Product launch", "mentions": 482, "sentiment": "positive"},
            {"topic": "Customer support", "mentions": 339, "sentiment": "mixed"},
            {"topic": "Pricing", "mentions": 271, "sentiment": "neutral"},
            {"topic": "Competitor comparison", "mentions": 198, "sentiment": "positive"},
        ],
        "hashtags": ["#BrandTracking", "#SocialListening", "#LaunchWeek", "#CustomerVoice"],
        "links": [
            {"title": "Launch announcement", "url": "brand.example/launch", "mentions": 144},
            {"title": "Industry comparison", "url": "media.example/ranking", "mentions": 91},
            {"title": "Community discussion", "url": "forum.example/thread", "mentions": 73},
        ],
        "aiSummary": (
            f"{brand} is seeing a sustained lift in conversation volume, led by launch coverage "
            "and comparison posts. Positive sentiment is strongest in news and creator channels, "
            "while support-related discussion is the main watch item."
        ),
        "insights": [
            "Conversation spikes cluster around weekday mornings, suggesting PR and analyst pickup.",
            "News domains create the largest reach multiplier even though social produces more raw mentions.",
            "Pricing questions are frequent but mostly neutral, making them a good content opportunity.",
        ],
        "hotHours": [
            {"hour": "08:00", "mentions": 146},
            {"hour": "11:00", "mentions": 218},
            {"hour": "14:00", "mentions": 172},
            {"hour": "19:00", "mentions": 133},
        ],
    }


def _normalize_series(mapping: Dict[str, Any], value_key: str) -> List[Dict[str, Any]]:
    return [{"date": day, value_key: value} for day, value in sorted(mapping.items())]


def _envelope_payload(response: Dict[str, Any]) -> Dict[str, Any]:
    payload = response.get("message", response.get("data", {}))
    return payload if isinstance(payload, dict) else {}


def _brand24_credentials() -> Tuple[str, str]:
    api_key = os.getenv("BRAND24_API_KEY")
    account_id = os.getenv("BRAND24_ACCOUNT_ID")
    if not api_key or not account_id:
        raise HTTPException(
            status_code=503,
            detail="Monitoring service credentials are not configured on the server.",
        )
    return api_key, account_id


def _normalize_projects_response(response: Dict[str, Any]) -> List[Dict[str, str]]:
    """Normalize every projects-list shape observed in the docs and live API."""
    payload: Any = response.get("data", response.get("message", []))
    if isinstance(payload, dict) and "projects_list" in payload:
        payload = payload["projects_list"]

    projects: List[Dict[str, str]] = []
    if isinstance(payload, dict):
        for raw_id, raw_project in payload.items():
            if isinstance(raw_project, str):
                projects.append({"id": str(raw_id), "name": raw_project})
                continue
            if not isinstance(raw_project, dict):
                continue
            project_id = raw_project.get("projectId") or raw_project.get("project_id") or raw_project.get("id") or raw_id
            project_name = raw_project.get("projectName") or raw_project.get("project_name") or raw_project.get("name")
            if project_id and project_name:
                projects.append({"id": str(project_id), "name": str(project_name)})

    if isinstance(payload, list):
        for item in payload:
            if not isinstance(item, dict):
                continue
            project_id = item.get("projectId") or item.get("project_id") or item.get("id")
            project_name = item.get("projectName") or item.get("project_name") or item.get("name")
            if project_id and project_name:
                projects.append({"id": str(project_id), "name": str(project_name)})

    deduplicated = {project["id"]: project for project in projects}
    return sorted(deduplicated.values(), key=lambda project: project["name"].casefold())


def _brand24_projects(force_refresh: bool = False) -> List[Dict[str, str]]:
    global _projects_cache

    now = time()
    if not force_refresh and _projects_cache and now - _projects_cache[0] < PROJECT_CACHE_TTL_SECONDS:
        return _projects_cache[1]

    api_key, account_id = _brand24_credentials()
    response = _brand24_get(f"/api-data/v1/account/{account_id}/projects_list/", api_key)
    projects = _normalize_projects_response(response)
    _projects_cache = (now, projects)
    return projects


def _number(value: Any, default: float = 0) -> float:
    if isinstance(value, bool):
        return default
    if isinstance(value, (int, float)):
        return value
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _whole_number(value: Any, default: int = 0) -> int:
    return int(round(_number(value, default)))


def _sentiment_count(day: Dict[str, Any], label: str, mentions: int) -> int:
    explicit = day.get(f"{label}_mentions")
    if explicit is not None:
        return _whole_number(explicit)

    sentiment = day.get("sentiment")
    if not isinstance(sentiment, dict):
        return mentions if label == "neutral" else 0

    value = _number(sentiment.get(label))
    return _whole_number(value * mentions if 0 <= value <= 1 else value)


def _normalize_daily_metrics(payload: Dict[str, Any], expected_dates: List[str]) -> Dict[str, Any]:
    raw_days = payload.get("days") if isinstance(payload.get("days"), list) else []
    days_by_date = {
        str(day.get("date")): day
        for day in raw_days
        if isinstance(day, dict) and day.get("date")
    }

    normalized_days: List[Dict[str, Any]] = []
    source_totals: Dict[str, Dict[str, Any]] = {}
    for day_name in expected_dates:
        raw_day = days_by_date.get(day_name, {})
        mentions = _whole_number(raw_day.get("mentions_count"))
        reach = _whole_number(raw_day.get("reach_total"))
        engagement = raw_day.get("engagement") if isinstance(raw_day.get("engagement"), dict) else {}
        likes = _whole_number(engagement.get("likes", raw_day.get("likes_count")))
        comments = _whole_number(engagement.get("comments", raw_day.get("comments_count")))
        shares = _whole_number(engagement.get("shares", raw_day.get("shares_count")))

        sources: List[Dict[str, Any]] = []
        raw_sources = raw_day.get("by_source") if isinstance(raw_day.get("by_source"), list) else []
        for raw_source in raw_sources:
            if not isinstance(raw_source, dict) or not raw_source.get("source"):
                continue
            source = str(raw_source["source"])
            source_mentions = _whole_number(raw_source.get("mentions_count"))
            source_reach = _whole_number(raw_source.get("reach"))
            sources.append({"source": source, "mentions": source_mentions, "reach": source_reach})
            aggregate = source_totals.setdefault(source, {"source": source, "mentions": 0, "reach": 0})
            aggregate["mentions"] += source_mentions
            aggregate["reach"] += source_reach

        normalized_days.append(
            {
                "date": day_name,
                "mentions": mentions,
                "reach": reach,
                "sentiment": {
                    "positive": _sentiment_count(raw_day, "positive", mentions),
                    "neutral": _sentiment_count(raw_day, "neutral", mentions),
                    "negative": _sentiment_count(raw_day, "negative", mentions),
                },
                "engagement": {"likes": likes, "comments": comments, "shares": shares},
                "sources": sources,
            }
        )

    sentiment_totals = {
        label: sum(day["sentiment"][label] for day in normalized_days)
        for label in ("positive", "neutral", "negative")
    }
    engagement_totals = {
        label: sum(day["engagement"][label] for day in normalized_days)
        for label in ("likes", "comments", "shares")
    }
    return {
        "days": normalized_days,
        "totals": {
            "mentions": sum(day["mentions"] for day in normalized_days),
            "reach": sum(day["reach"] for day in normalized_days),
            "sentiment": sentiment_totals,
            "engagement": engagement_totals,
        },
        "sources": sorted(source_totals.values(), key=lambda source: source["mentions"], reverse=True),
    }


def _feature_state(items: List[Any], upstream_status: Optional[str] = None) -> str:
    if upstream_status == "unavailable":
        return "unavailable"
    return "ready" if items else "empty"


def _normalize_topics(payload: Dict[str, Any]) -> List[Dict[str, Any]]:
    raw_topics = payload.get("topics") if isinstance(payload.get("topics"), list) else []
    topics: List[Dict[str, Any]] = []
    for index, item in enumerate(raw_topics):
        if not isinstance(item, dict):
            continue

        raw_sentiment = item.get("sentiment") if isinstance(item.get("sentiment"), dict) else {}
        sentiment = {
            label: _number(raw_sentiment.get(label))
            for label in ("positive", "neutral", "negative")
        }
        # Live responses use 0–1 shares; some documented examples use 0–100.
        if max(sentiment.values(), default=0) <= 1:
            sentiment = {label: value * 100 for label, value in sentiment.items()}
        dominant_sentiment = max(sentiment, key=sentiment.get) if any(sentiment.values()) else "neutral"

        topics.append(
            {
                "id": str(item.get("topic_id", index)),
                "name": str(item.get("topic_name") or "Untitled topic"),
                "description": str(item.get("description") or ""),
                "mentions": _whole_number(item.get("mentions")),
                "reach": _whole_number(item.get("reach")),
                "sentiment": sentiment,
                "dominantSentiment": dominant_sentiment,
                "shareOfVoice": _number(item.get("share_of_voice")),
            }
        )
    return topics


def _normalize_insights(payload: Dict[str, Any]) -> List[Dict[str, str]]:
    raw_insights = payload.get("insights")
    normalized: List[Dict[str, str]] = []
    if isinstance(raw_insights, list):
        for index, item in enumerate(raw_insights):
            if isinstance(item, str) and item.strip():
                normalized.append({"title": f"Insight {index + 1}", "text": item.strip()})
            elif isinstance(item, dict):
                title = item.get("headline") or item.get("title") or item.get("insightType") or f"Insight {index + 1}"
                body = item.get("text") or item.get("description") or item.get("recommendation")
                if body:
                    normalized.append({"title": str(title), "text": str(body)})
    else:
        for key, title in (
            ("headline", "Headline"),
            ("trends", "Trend"),
            ("insights", "Insight"),
            ("recommendations", "Recommendation"),
        ):
            value = payload.get(key)
            if isinstance(value, str) and value.strip():
                normalized.append({"title": title, "text": value.strip()})
    return normalized


def _briefing_feature_call(
    feature: str,
    path: str,
    api_key: str,
    query: Dict[str, Any],
) -> Tuple[str, Dict[str, Any], Optional[Dict[str, Any]]]:
    try:
        return feature, _envelope_payload(_brand24_get(path, api_key, query)), None
    except HTTPException as error:
        return feature, {}, {
            "feature": feature,
            "status": error.status_code,
            "message": str(error.detail)[:300],
        }


def _live_briefing(project: Dict[str, str], date_from: date, date_to: date) -> Dict[str, Any]:
    api_key, _ = _brand24_credentials()
    days = (date_to - date_from).days + 1
    dates = [(date_from + timedelta(days=index)).isoformat() for index in range(days)]
    previous_to = date_from - timedelta(days=1)
    previous_from = previous_to - timedelta(days=days - 1)
    previous_dates = [
        (previous_from + timedelta(days=index)).isoformat()
        for index in range(days)
    ]
    date_query = {"date_from": dates[0], "date_to": dates[-1]}
    requests = {
        "metrics": (
            f"/api-data/v1/project/{project['id']}/daily-metrics",
            {"from": dates[0], "to": dates[-1], "includeBySource": "true"},
        ),
        "comparison": (
            f"/api-data/v1/project/{project['id']}/daily-metrics",
            {"from": previous_dates[0], "to": previous_dates[-1], "includeBySource": "true"},
        ),
        "summary": (f"/api-data/v1/project/{project['id']}/ai-summary", date_query),
        "events": (f"/api-data/v1/project/{project['id']}/project_events", date_query),
        "topics": (f"/api-data/v1/project/{project['id']}/topics", date_query),
        "insights": (f"/api-data/v1/project/{project['id']}/ai-insights", date_query),
        "coverage": (
            f"/api-data/v1/project/{project['id']}/mentions",
            {"date_from": dates[0], "date_to": dates[-1], "limit": 8},
        ),
    }

    payloads: Dict[str, Dict[str, Any]] = {}
    warnings: List[Dict[str, Any]] = []
    with ThreadPoolExecutor(max_workers=len(requests)) as pool:
        futures = [
            pool.submit(_briefing_feature_call, feature, path, api_key, query)
            for feature, (path, query) in requests.items()
        ]
        for future in as_completed(futures):
            feature, payload, warning = future.result()
            payloads[feature] = payload
            if warning:
                warnings.append(warning)

    if not payloads.get("metrics"):
        metrics_warning = next((warning for warning in warnings if warning["feature"] == "metrics"), None)
        raise HTTPException(
            status_code=502,
            detail=metrics_warning["message"] if metrics_warning else "Daily monitoring metrics are unavailable.",
        )

    metrics = _normalize_daily_metrics(payloads["metrics"], dates)
    previous_metrics = (
        _normalize_daily_metrics(payloads["comparison"], previous_dates)
        if payloads.get("comparison")
        else None
    )

    def percent_change(current: float, previous: float) -> Optional[float]:
        if not previous:
            return 0.0 if not current else None
        return round(((current - previous) / previous) * 100, 1)

    comparison: Dict[str, Any] = {
        "status": "ready" if previous_metrics else "unavailable",
        "dateRange": {"from": previous_dates[0], "to": previous_dates[-1]},
        "changes": {},
    }
    if previous_metrics:
        current_totals = metrics["totals"]
        previous_totals = previous_metrics["totals"]
        current_engagement = sum(current_totals["engagement"].values())
        previous_engagement = sum(previous_totals["engagement"].values())
        current_sentiment_total = sum(current_totals["sentiment"].values())
        previous_sentiment_total = sum(previous_totals["sentiment"].values())
        current_negative_share = (
            current_totals["sentiment"]["negative"] / current_sentiment_total * 100
            if current_sentiment_total
            else 0
        )
        previous_negative_share = (
            previous_totals["sentiment"]["negative"] / previous_sentiment_total * 100
            if previous_sentiment_total
            else 0
        )
        comparison["previousTotals"] = previous_totals
        comparison["changes"] = {
            "mentions": percent_change(current_totals["mentions"], previous_totals["mentions"]),
            "reach": percent_change(current_totals["reach"], previous_totals["reach"]),
            "engagement": percent_change(current_engagement, previous_engagement),
            "negativeSharePoints": round(current_negative_share - previous_negative_share, 1),
        }
    summary_text = str(payloads.get("summary", {}).get("summary") or "").strip()
    summary_text = summary_text.replace("<br><br>", "\n\n").replace("<br>", "\n")

    event_payload = payloads.get("events", {})
    raw_events = event_payload.get("anomalies") if isinstance(event_payload.get("anomalies"), list) else []
    events = [
        {
            "date": str(item.get("anomaly_date", "")),
            "description": str(item.get("description", "")),
            "mentions": _whole_number(item.get("peak_mentions")),
            "reach": _whole_number(item.get("peak_reach")),
        }
        for item in raw_events
        if isinstance(item, dict)
    ]

    topic_payload = payloads.get("topics", {})
    topics = _normalize_topics(topic_payload)

    insights = _normalize_insights(payloads.get("insights", {}))
    coverage_payload = payloads.get("coverage", {})
    raw_coverage = coverage_payload.get("results") if isinstance(coverage_payload.get("results"), list) else []
    coverage = [
        _normalize_mention(item)
        for item in raw_coverage
        if isinstance(item, dict)
    ]
    coverage = sorted(
        coverage,
        key=lambda item: (
            item["restricted"],
            not bool(item["title"] or item["content"]),
            item["sentiment"] == "unknown",
        ),
    )[:5]
    return {
        "project": project,
        "dateRange": {"from": dates[0], "to": dates[-1], "days": days},
        "metrics": metrics,
        "comparison": comparison,
        "summary": {"status": "ready" if summary_text else "empty", "text": summary_text},
        "events": {"status": _feature_state(events), "items": events},
        "topics": {
            "status": _feature_state(topics, str(topic_payload.get("status") or "")),
            "items": topics,
        },
        "insights": {"status": _feature_state(insights), "items": insights},
        "coverage": {"status": _feature_state(coverage), "items": coverage},
        "warnings": sorted(warnings, key=lambda warning: warning["feature"]),
    }


def _brand24_mention_categories(force_refresh: bool = False) -> List[str]:
    global _categories_cache

    now = time()
    if not force_refresh and _categories_cache and now - _categories_cache[0] < 3600:
        return _categories_cache[1]

    api_key, _ = _brand24_credentials()
    payload = _envelope_payload(_brand24_get("/api-data/v1/mentions/categories", api_key))
    raw_categories = payload.get("categories") if isinstance(payload.get("categories"), list) else []
    categories = sorted({str(category).strip().lower() for category in raw_categories if str(category).strip()})
    _categories_cache = (now, categories)
    return categories


def _parse_date_parameter(value: Optional[str], name: str) -> Optional[date]:
    if value is None:
        return None
    try:
        return date.fromisoformat(value)
    except ValueError:
        raise HTTPException(status_code=400, detail=f"{name} must use YYYY-MM-DD format.")


def _normalize_mention(item: Dict[str, Any]) -> Dict[str, Any]:
    published_date = str(item.get("date") or "")
    published_time = str(item.get("time") or "")
    title = str(item["title"]).strip() if item.get("title") else ""
    content = str(item["content"]).strip() if item.get("content") else ""
    source = str(item["source"]).strip() if item.get("source") else ""
    host = str(item["host"]).strip() if item.get("host") else ""
    category = str(item["category"]).strip() if item.get("category") else "Unknown"
    category_token = category.casefold()

    raw_sentiment = item.get("sentiment")
    sentiment = {-1: "negative", 0: "neutral", 1: "positive"}.get(raw_sentiment)
    if sentiment is None and isinstance(raw_sentiment, str) and raw_sentiment.casefold() in {"positive", "neutral", "negative"}:
        sentiment = raw_sentiment.casefold()
    sentiment = sentiment or "unknown"

    source_url = source if source.startswith(("https://", "http://")) else None
    restricted_platform = category_token in {"facebook", "instagram", "x", "x (twitter)", "twitter"}
    restricted = restricted_platform and not (title or content)
    if restricted:
        if "x" in category_token or "twitter" in category_token:
            restriction_reason = "Post text is unavailable through the API for X/Twitter mentions."
        else:
            restriction_reason = f"Post text and source links are unavailable through the API for {category}."
    else:
        restriction_reason = None

    raw_tags = item.get("tags") if isinstance(item.get("tags"), list) else []
    tags = [str(tag) for tag in raw_tags if str(tag).strip()]
    identity = "|".join((published_date, published_time, source, host, title, content[:80]))
    return {
        "id": sha256(identity.encode("utf-8")).hexdigest()[:20],
        "date": published_date,
        "time": published_time,
        "title": title or None,
        "content": content or None,
        "source": source or None,
        "sourceUrl": source_url,
        "host": host or None,
        "category": category,
        "sentiment": sentiment,
        "tags": tags,
        "restricted": restricted,
        "restrictionReason": restriction_reason,
    }


def _live_mentions(
    project: Dict[str, str],
    date_from: date,
    date_to: date,
    limit: int,
    cursor: Optional[str],
    sentiment: Optional[str],
    category: Optional[str],
) -> Dict[str, Any]:
    api_key, _ = _brand24_credentials()
    query: Dict[str, Any] = {
        "date_from": date_from.isoformat(),
        "date_to": date_to.isoformat(),
        "limit": limit,
    }
    if sentiment:
        query["sentiment"] = sentiment
    if category:
        query["category"] = category

    raw_results: List[Dict[str, Any]] = []
    next_cursor = cursor
    has_more = False
    seen_cursors = {cursor} if cursor else set()
    # Brand24 may apply filters after forming an internal page, yielding an empty
    # result with has_more_mentions=true. Advance through a bounded number of
    # sparse pages so one app-level page contains useful filtered evidence.
    for _ in range(8):
        page_query = {**query}
        if next_cursor:
            page_query["cursor"] = next_cursor
        payload = _envelope_payload(
            _brand24_get(f"/api-data/v1/project/{project['id']}/mentions", api_key, page_query)
        )
        page_results = payload.get("results") if isinstance(payload.get("results"), list) else []
        raw_results.extend(item for item in page_results if isinstance(item, dict))
        has_more = bool(payload.get("has_more_mentions"))
        candidate_cursor = str(payload["cursor"]) if payload.get("cursor") else None
        if len(raw_results) >= limit or not has_more or not candidate_cursor or candidate_cursor in seen_cursors:
            next_cursor = candidate_cursor
            break
        seen_cursors.add(candidate_cursor)
        next_cursor = candidate_cursor

    normalized_items = [_normalize_mention(item) for item in raw_results]
    deduplicated_items = list({item["id"]: item for item in normalized_items}.values())
    return {
        "project": project,
        "dateRange": {"from": date_from.isoformat(), "to": date_to.isoformat()},
        "filters": {"sentiment": sentiment, "category": category},
        "items": deduplicated_items,
        "pagination": {
            "hasMore": has_more,
            "cursor": next_cursor,
        },
    }


def _safe_url(value: Any) -> Optional[str]:
    candidate = str(value or "").strip()
    return candidate if candidate.startswith(("https://", "http://")) else None


def _live_sources(project: Dict[str, str], date_from: date, date_to: date) -> Dict[str, Any]:
    api_key, _ = _brand24_credentials()
    query = {"date_from": date_from.isoformat(), "date_to": date_to.isoformat()}
    requests = {
        "domains": (f"/api-data/v1/project/{project['id']}/domains/", query),
        "activeSites": (f"/api-data/v1/project/{project['id']}/most-active-sites", query),
        "links": (f"/api-data/v1/project/{project['id']}/trending-links", query),
        "hashtags": (f"/api-data/v1/project/{project['id']}/trending-hashtags", query),
        "authors": (f"/api-data/v1/project/{project['id']}/most-followers", query),
        "hotHours": (f"/api-data/v1/project/{project['id']}/hot-hours", query),
    }

    payloads: Dict[str, Dict[str, Any]] = {}
    warnings: List[Dict[str, Any]] = []
    with ThreadPoolExecutor(max_workers=len(requests)) as pool:
        futures = [
            pool.submit(_briefing_feature_call, feature, path, api_key, feature_query)
            for feature, (path, feature_query) in requests.items()
        ]
        for future in as_completed(futures):
            feature, payload, warning = future.result()
            payloads[feature] = payload
            if warning:
                warnings.append(warning)

    domain_payload = payloads.get("domains", {})
    raw_domains = domain_payload.get("domains") if isinstance(domain_payload.get("domains"), list) else []
    domains = [
        {
            "domain": str(item.get("domain") or "Unknown"),
            "mentions": _whole_number(item.get("mentions_count")),
            "reach": _whole_number(item.get("reach")),
            "visits": _whole_number(item.get("visits")),
            "influenceScore": _whole_number(item.get("influence_score")),
        }
        for item in raw_domains
        if isinstance(item, dict)
    ]

    active_payload = payloads.get("activeSites", {})
    raw_active_sites = active_payload.get("sites") if isinstance(active_payload.get("sites"), list) else []
    active_sites = [
        {
            "domain": str(item.get("domain") or "Unknown"),
            "mentions": _whole_number(item.get("mentions_count")),
            "reach": _whole_number(item.get("reach")),
        }
        for item in raw_active_sites
        if isinstance(item, dict)
    ]

    link_payload = payloads.get("links", {})
    raw_links = link_payload.get("trending_links") if isinstance(link_payload.get("trending_links"), list) else []
    links = [
        {"url": url, "mentions": _whole_number(item.get("mentions_count"))}
        for item in raw_links
        if isinstance(item, dict) and (url := _safe_url(item.get("url")))
    ]

    hashtag_payload = payloads.get("hashtags", {})
    raw_hashtags = hashtag_payload.get("hashtags") if isinstance(hashtag_payload.get("hashtags"), list) else []
    hashtags = [
        {
            "hashtag": str(item.get("hashtag") or ""),
            "mentions": _whole_number(item.get("mentions_count")),
            "reach": _whole_number(item.get("social_media_reach")),
            "sentimentScore": _number(item.get("sentiment_score")) if item.get("sentiment_score") is not None else None,
        }
        for item in raw_hashtags
        if isinstance(item, dict) and item.get("hashtag")
    ]

    author_payload = payloads.get("authors", {})
    raw_authors = author_payload.get("authors") if isinstance(author_payload.get("authors"), list) else []
    authors = [
        {
            "name": str(item.get("name") or "Unknown author"),
            "url": _safe_url(item.get("url")),
            "followers": _whole_number(item.get("followers_count")),
            "mentions": _whole_number(item.get("mentions_count")),
            "reach": _whole_number(item.get("reach")),
        }
        for item in raw_authors
        if isinstance(item, dict)
    ]

    hour_payload = payloads.get("hotHours", {})
    raw_hot_hours = hour_payload.get("hot_hours") if isinstance(hour_payload.get("hot_hours"), list) else []
    day_names = {1: "Monday", 2: "Tuesday", 3: "Wednesday", 4: "Thursday", 5: "Friday", 6: "Saturday", 7: "Sunday"}
    hot_hours = [
        {
            "day": _whole_number(item.get("day_of_week")),
            "dayName": day_names.get(_whole_number(item.get("day_of_week")), "Unknown"),
            "hour": _whole_number(item.get("hour")),
            "mentions": _number(item.get("mentions_count")),
        }
        for item in raw_hot_hours
        if isinstance(item, dict)
    ]
    hot_hours.sort(key=lambda item: item["mentions"], reverse=True)

    peak_hour = hot_hours[0] if hot_hours else None
    return {
        "project": project,
        "dateRange": {"from": date_from.isoformat(), "to": date_to.isoformat()},
        "summary": {
            "totalDomains": _whole_number(domain_payload.get("total_domains", len(domains))),
            "totalLinks": _whole_number(link_payload.get("total_links", len(links))),
            "totalAuthors": _whole_number(author_payload.get("total_authors", len(authors))),
            "peakHour": peak_hour,
        },
        "domains": {"status": _feature_state(domains), "items": domains},
        "activeSites": {"status": _feature_state(active_sites), "items": active_sites},
        "links": {"status": _feature_state(links), "items": links},
        "hashtags": {"status": _feature_state(hashtags), "items": hashtags},
        "authors": {"status": _feature_state(authors), "items": authors},
        "hotHours": {"status": _feature_state(hot_hours), "items": hot_hours},
        "warnings": sorted(warnings, key=lambda warning: warning["feature"]),
    }


def _live_topics(project: Dict[str, str], date_from: date, date_to: date) -> Dict[str, Any]:
    api_key, _ = _brand24_credentials()
    payload = _envelope_payload(
        _brand24_get(
            f"/api-data/v1/project/{project['id']}/topics",
            api_key,
            {"date_from": date_from.isoformat(), "date_to": date_to.isoformat()},
        )
    )
    topics = _normalize_topics(payload)
    topics.sort(key=lambda topic: topic["shareOfVoice"], reverse=True)
    leading_topic = topics[0] if topics else None
    most_negative_topic = max(topics, key=lambda topic: topic["sentiment"]["negative"]) if topics else None

    return {
        "project": project,
        "dateRange": {"from": date_from.isoformat(), "to": date_to.isoformat()},
        "status": _feature_state(topics, str(payload.get("status") or "")),
        "summary": {
            "topicCount": len(topics),
            "topicMentions": sum(topic["mentions"] for topic in topics),
            "topicReach": sum(topic["reach"] for topic in topics),
            "leadingTopic": leading_topic["name"] if leading_topic else None,
            "mostNegativeTopic": most_negative_topic["name"] if most_negative_topic else None,
        },
        "items": topics,
    }


def _export_all_mentions(
    project: Dict[str, str], date_from: date, date_to: date, max_pages: int = 20
) -> Dict[str, Any]:
    api_key, _ = _brand24_credentials()
    cursor: Optional[str] = None
    seen_cursors = set()
    items: List[Dict[str, Any]] = []
    has_more = False

    for _ in range(max_pages):
        query: Dict[str, Any] = {
            "date_from": date_from.isoformat(),
            "date_to": date_to.isoformat(),
            "limit": 500,
        }
        if cursor:
            query["cursor"] = cursor
        payload = _envelope_payload(
            _brand24_get(f"/api-data/v1/project/{project['id']}/mentions", api_key, query)
        )
        raw_results = payload.get("results") if isinstance(payload.get("results"), list) else []
        items.extend(_normalize_mention(item) for item in raw_results if isinstance(item, dict))
        has_more = bool(payload.get("has_more_mentions"))
        next_cursor = str(payload["cursor"]) if payload.get("cursor") else None
        if not has_more or not next_cursor or next_cursor in seen_cursors:
            break
        seen_cursors.add(next_cursor)
        cursor = next_cursor

    deduplicated = list({item["id"]: item for item in items}.values())
    return {"items": deduplicated, "truncated": has_more, "limit": max_pages * 500}


def _build_export_dataset(
    project: Dict[str, str], date_from: date, date_to: date, sections: List[str]
) -> Dict[str, Any]:
    section_set = set(sections)
    dates = [
        (date_from + timedelta(days=index)).isoformat()
        for index in range((date_to - date_from).days + 1)
    ]
    api_key, _ = _brand24_credentials()

    def load_metrics() -> Dict[str, Any]:
        payload = _envelope_payload(
            _brand24_get(
                f"/api-data/v1/project/{project['id']}/daily-metrics",
                api_key,
                {"from": date_from.isoformat(), "to": date_to.isoformat(), "includeBySource": "true"},
            )
        )
        if not payload:
            raise HTTPException(status_code=502, detail="Daily monitoring metrics are unavailable for this export.")
        return _normalize_daily_metrics(payload, dates)

    loaders: Dict[str, Callable[[], Dict[str, Any]]] = {
        "metrics": load_metrics,
    }
    if "mentions" in section_set:
        loaders["mentions"] = lambda: _export_all_mentions(project, date_from, date_to)
    if section_set.intersection({"sources", "authors", "links", "hashtags"}):
        loaders["sources"] = lambda: _live_sources(project, date_from, date_to)
    if "topics" in section_set:
        loaders["topics"] = lambda: _live_topics(project, date_from, date_to)

    results: Dict[str, Any] = {}
    with ThreadPoolExecutor(max_workers=len(loaders)) as pool:
        futures = {pool.submit(loader): name for name, loader in loaders.items()}
        for future in as_completed(futures):
            results[futures[future]] = future.result()

    return {
        "project": project,
        "dateRange": {"from": date_from.isoformat(), "to": date_to.isoformat()},
        "generatedAt": int(time()),
        **results,
    }


_EXPORT_ORANGE = "ED6E1F"
_EXPORT_ORANGE_DARK = "B83A18"
_EXPORT_ORANGE_SOFT = "FFF1E6"
_EXPORT_GOLD = "ECA81A"
_EXPORT_TEXT = "24150E"
_EXPORT_MUTED = "7C6256"
_EXPORT_LINE = "ECD6C7"


def _excel_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, str):
        value = value[:32767]
        if value.startswith(("=", "+", "-", "@")):
            return f"'{value}"
    return value


def _style_export_sheet(sheet: Any, freeze_panes: str = "A2") -> None:
    sheet.freeze_panes = freeze_panes
    sheet.sheet_view.showGridLines = False
    sheet.auto_filter.ref = sheet.dimensions
    for column_cells in sheet.columns:
        letter = get_column_letter(column_cells[0].column)
        width = max((len(str(cell.value or "")) for cell in column_cells), default=8) + 2
        sheet.column_dimensions[letter].width = min(max(width, 11), 48)
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor=_EXPORT_ORANGE_DARK)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 24


def _write_export_table(sheet: Any, headers: List[str], rows: List[List[Any]]) -> None:
    sheet.append(headers)
    for row in rows:
        sheet.append([_excel_value(value) for value in row])
    _style_export_sheet(sheet)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=Side(style="hair", color=_EXPORT_LINE))


def _build_export_workbook(dataset: Dict[str, Any], payload: ExportCreate) -> bytes:
    translations = {
        "en": {
            "report": "Media Intelligence Report",
            "overview": "Overview",
            "daily": "Daily Metrics",
            "mentions": "Mentions",
            "sources": "Sources",
            "topics": "Topics",
            "authors": "Authors",
            "links": "Links",
            "hashtags": "Hashtags",
        },
        "ms": {
            "report": "Laporan Risikan Media",
            "overview": "Ringkasan",
            "daily": "Metrik Harian",
            "mentions": "Sebutan",
            "sources": "Sumber",
            "topics": "Topik",
            "authors": "Pengarang",
            "links": "Pautan",
            "hashtags": "Tanda Pagar",
        },
    }
    labels = translations.get(payload.language, translations["en"])
    malay_terms = {
        "Project": "Projek", "Period": "Tempoh", "Organization": "Organisasi",
        "Total mentions": "Jumlah sebutan", "Mentions": "Sebutan", "Engagement": "Penglibatan", "Estimated reach": "Anggaran capaian",
        "Total engagement": "Jumlah penglibatan", "Peak engagement day": "Hari penglibatan tertinggi",
        "Sentiment": "Sentimen", "Positive": "Positif", "Neutral": "Neutral", "Negative": "Negatif",
        "Leading source": "Sumber utama", "Source": "Sumber", "Reach": "Capaian",
        "Daily conversation": "Perbualan harian", "Count": "Bilangan", "Date": "Tarikh",
        "Sentiment mix": "Pecahan sentimen", "Methodology note": "Nota metodologi",
        "Time": "Masa", "Title": "Tajuk", "Content": "Kandungan", "Source URL": "URL sumber",
        "Host": "Hos", "Category": "Kategori", "Tags": "Tag", "Restricted": "Terhad",
        "Restriction reason": "Sebab sekatan", "Likes": "Suka", "Comments": "Komen",
        "Shares": "Kongsi", "Domain": "Domain", "Monthly visits": "Lawatan bulanan",
        "Influence score": "Skor pengaruh", "Topic": "Topik", "Description": "Penerangan",
        "Share of voice %": "Bahagian suara %", "Positive %": "Positif %", "Neutral %": "Neutral %",
        "Negative %": "Negatif %", "Dominant sentiment": "Sentimen dominan", "Author": "Pengarang",
        "Profile URL": "URL profil", "Followers": "Pengikut", "Estimated reach": "Anggaran capaian",
        "URL": "URL", "Hashtag": "Tanda pagar", "Social media reach": "Capaian media sosial",
        "Sentiment score": "Skor sentimen", "Yes": "Ya", "No": "Tidak",
    }

    def translate(value: str) -> str:
        return malay_terms.get(value, value) if payload.language == "ms" else value

    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.properties.title = payload.reportTitle or labels["report"]
    workbook.properties.subject = f"Media monitoring export for {dataset['project']['name']}"
    workbook.properties.creator = "Zestar Media Intelligence"

    metrics = dataset["metrics"]
    totals = metrics["totals"]
    total_engagement = sum(totals["engagement"].values())
    peak_day = max(
        metrics["days"],
        key=lambda item: sum(item["engagement"].values()),
        default=None,
    )

    if "overview" in payload.sections:
        sheet = workbook.create_sheet(labels["overview"])
        sheet.sheet_view.showGridLines = False
        sheet.merge_cells("A1:I2")
        title_cell = sheet["A1"]
        title_cell.value = payload.reportTitle or labels["report"]
        title_cell.fill = PatternFill("solid", fgColor=_EXPORT_ORANGE_DARK)
        title_cell.font = Font(color="FFFFFF", bold=True, size=22)
        title_cell.alignment = Alignment(vertical="center")
        for row in sheet["A1:I2"]:
            for cell in row:
                cell.fill = PatternFill("solid", fgColor=_EXPORT_ORANGE_DARK)
        sheet["A3"] = translate("Project")
        sheet["B3"] = dataset["project"]["name"]
        sheet["D3"] = translate("Period")
        sheet["E3"] = f"{dataset['dateRange']['from']} to {dataset['dateRange']['to']}"
        sheet["G3"] = translate("Organization")
        sheet["H3"] = payload.organization or "—"

        kpis = [
            (translate("Total mentions"), totals["mentions"]),
            (translate("Estimated reach"), totals["reach"]),
            (translate("Total engagement"), total_engagement),
            (translate("Peak engagement day"), peak_day["date"] if peak_day else "—"),
        ]
        for column, (label, value) in zip(("A", "C", "E", "G"), kpis):
            sheet[f"{column}5"] = label
            sheet[f"{column}6"] = value
            sheet[f"{column}5"].font = Font(color=_EXPORT_MUTED, bold=True, size=9)
            sheet[f"{column}6"].font = Font(color=_EXPORT_ORANGE_DARK, bold=True, size=18)
            sheet.merge_cells(f"{column}5:{get_column_letter(sheet[column + '5'].column + 1)}5")
            sheet.merge_cells(f"{column}6:{get_column_letter(sheet[column + '6'].column + 1)}7")

        sheet["A9"] = translate("Sentiment")
        sheet["B9"] = translate("Mentions")
        for index, sentiment in enumerate(("positive", "neutral", "negative"), start=10):
            sheet.cell(index, 1, translate(sentiment.title()))
            sheet.cell(index, 2, totals["sentiment"][sentiment])
        sheet["D9"] = translate("Leading source")
        sheet["E9"] = translate("Mentions")
        sheet["F9"] = translate("Reach")
        for index, source in enumerate(metrics["sources"][:8], start=10):
            sheet.cell(index, 4, source["source"].title())
            sheet.cell(index, 5, source["mentions"])
            sheet.cell(index, 6, source["reach"])
        for cell in (*sheet[9][0:2], *sheet[9][3:6]):
            cell.fill = PatternFill("solid", fgColor=_EXPORT_ORANGE)
            cell.font = Font(color="FFFFFF", bold=True)

        chart_data = workbook.create_sheet("_ChartData")
        chart_data.append([translate("Date"), translate("Mentions"), translate("Engagement")])
        for item in metrics["days"]:
            chart_data.append([item["date"], item["mentions"], sum(item["engagement"].values())])
        chart_data.sheet_state = "hidden"

        line_chart = LineChart()
        line_chart.title = translate("Daily conversation")
        line_chart.y_axis.title = translate("Count")
        line_chart.x_axis.title = translate("Date")
        line_chart.add_data(Reference(chart_data, min_col=2, max_col=3, min_row=1, max_row=len(metrics["days"]) + 1), titles_from_data=True)
        line_chart.set_categories(Reference(chart_data, min_col=1, min_row=2, max_row=len(metrics["days"]) + 1))
        line_chart.height = 7
        line_chart.width = 14
        sheet.add_chart(line_chart, "A20")

        pie_chart = PieChart()
        pie_chart.title = translate("Sentiment mix")
        pie_chart.add_data(Reference(sheet, min_col=2, min_row=9, max_row=12), titles_from_data=True)
        pie_chart.set_categories(Reference(sheet, min_col=1, min_row=10, max_row=12))
        pie_chart.height = 7
        pie_chart.width = 9
        sheet.add_chart(pie_chart, "G9")
        for column in range(1, 10):
            sheet.column_dimensions[get_column_letter(column)].width = 16
        sheet["A36"] = translate("Methodology note")
        sheet["A37"] = (
            "Sebutan individu tidak mengandungi capaian atau penglibatan bagi setiap siaran. Sumber sosial terhad mungkin tidak menyertakan teks dan pautan siaran."
            if payload.language == "ms" else
            "Individual mentions do not include per-post reach or engagement. Restricted social sources may omit post text and links."
        )
        sheet["A36"].font = Font(bold=True, color=_EXPORT_ORANGE_DARK)
        sheet.merge_cells("A37:I38")
        sheet["A37"].alignment = Alignment(wrap_text=True, vertical="top")

    if "daily" in payload.sections:
        sheet = workbook.create_sheet(labels["daily"])
        _write_export_table(sheet, [translate(value) for value in ["Date", "Mentions", "Reach", "Positive", "Neutral", "Negative", "Likes", "Comments", "Shares", "Total engagement"]], [
            [item["date"], item["mentions"], item["reach"], item["sentiment"]["positive"], item["sentiment"]["neutral"], item["sentiment"]["negative"], item["engagement"]["likes"], item["engagement"]["comments"], item["engagement"]["shares"], sum(item["engagement"].values())]
            for item in metrics["days"]
        ])

    if "mentions" in payload.sections:
        sheet = workbook.create_sheet(labels["mentions"])
        mention_data = dataset.get("mentions", {})
        _write_export_table(sheet, [translate(value) for value in ["Date", "Time", "Title", "Content", "Source", "Source URL", "Host", "Category", "Sentiment", "Tags", "Restricted", "Restriction reason"]], [
            [item["date"], item["time"], item["title"], item["content"], item["source"], item["sourceUrl"], item["host"], item["category"], item["sentiment"], ", ".join(item["tags"]), translate("Yes") if item["restricted"] else translate("No"), item["restrictionReason"]]
            for item in mention_data.get("items", [])
        ])
        if mention_data.get("truncated"):
            sheet.insert_rows(1)
            sheet["A1"] = f"Export capped at {mention_data['limit']:,} mentions. Narrow the reporting period for a complete raw export."

    source_data = dataset.get("sources", {})
    if "sources" in payload.sections:
        sheet = workbook.create_sheet(labels["sources"])
        _write_export_table(sheet, [translate(value) for value in ["Domain", "Mentions", "Reach", "Monthly visits", "Influence score"]], [
            [item["domain"], item["mentions"], item["reach"], item["visits"], item["influenceScore"]]
            for item in source_data.get("domains", {}).get("items", [])
        ])
    if "topics" in payload.sections:
        sheet = workbook.create_sheet(labels["topics"])
        _write_export_table(sheet, [translate(value) for value in ["Topic", "Description", "Mentions", "Reach", "Share of voice %", "Positive %", "Neutral %", "Negative %", "Dominant sentiment"]], [
            [item["name"], item["description"], item["mentions"], item["reach"], item["shareOfVoice"], item["sentiment"]["positive"], item["sentiment"]["neutral"], item["sentiment"]["negative"], item["dominantSentiment"]]
            for item in dataset.get("topics", {}).get("items", [])
        ])
    if "authors" in payload.sections:
        sheet = workbook.create_sheet(labels["authors"])
        _write_export_table(sheet, [translate(value) for value in ["Author", "Profile URL", "Followers", "Mentions", "Estimated reach"]], [
            [item["name"], item["url"], item["followers"], item["mentions"], item["reach"]]
            for item in source_data.get("authors", {}).get("items", [])
        ])
    if "links" in payload.sections:
        sheet = workbook.create_sheet(labels["links"])
        _write_export_table(sheet, [translate(value) for value in ["URL", "Mentions"]], [
            [item["url"], item["mentions"]] for item in source_data.get("links", {}).get("items", [])
        ])
    if "hashtags" in payload.sections:
        sheet = workbook.create_sheet(labels["hashtags"])
        _write_export_table(sheet, [translate(value) for value in ["Hashtag", "Mentions", "Social media reach", "Sentiment score"]], [
            [item["hashtag"], item["mentions"], item["reach"], item["sentimentScore"]]
            for item in source_data.get("hashtags", {}).get("items", [])
        ])

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _build_export_presentation(dataset: Dict[str, Any], payload: ExportCreate) -> bytes:
    return build_pptx_report(
        dataset,
        {
            "language": payload.language,
            "reportTitle": payload.reportTitle,
            "organization": payload.organization,
            "sections": payload.sections,
        },
    )


def _live_dashboard(project_id: str, api_key: str, brand: str, days: int) -> Dict[str, Any]:
    dates = _date_range(days)
    params = {"date_from": dates[0], "date_to": dates[-1]}

    mentions_count = _brand24_get(
        f"/api-data/v1/project/{project_id}/mentions/count",
        api_key,
        params,
    )
    sentiment = _brand24_get(
        f"/api-data/v1/project/{project_id}/mentions/sentiment",
        api_key,
        params,
    )
    reach = _brand24_get(
        f"/api-data/v1/project/{project_id}/mentions/reach",
        api_key,
        params,
    )
    topics = _brand24_get(f"/api-data/v1/project/{project_id}/topics", api_key, params)
    summary = _brand24_get(f"/api-data/v1/project/{project_id}/ai-summary", api_key, params)
    insights = _brand24_get(f"/api-data/v1/project/{project_id}/ai-insights", api_key, params)

    count_data = _envelope_payload(mentions_count)
    reach_data = _envelope_payload(reach)
    sentiment_data = _envelope_payload(sentiment)
    topics_data = _envelope_payload(topics)
    summary_data = _envelope_payload(summary)
    insights_data = _envelope_payload(insights)

    mention_counts = count_data.get("mentions_count", {})
    mentions_trend = [{"date": day, "mentions": mention_counts.get(day, 0)} for day in dates]

    social_reach = reach_data.get("social_media_reach", {})
    non_social_reach = reach_data.get("non_social_media_reach", {})
    reach_trend = [
        {
            "date": day,
            "reach": social_reach.get(day, 0) + non_social_reach.get(day, 0),
        }
        for day in dates
    ]

    positive_mentions = sentiment_data.get("positive_mentions", {})
    negative_mentions = sentiment_data.get("negative_mentions", {})
    total_mentions = sentiment_data.get("mentions", mention_counts)
    sentiment_trend = []
    for day in dates:
        positive_day = positive_mentions.get(day, 0)
        negative_day = negative_mentions.get(day, 0)
        total_day = total_mentions.get(day, 0)
        sentiment_trend.append(
            {
                "date": day,
                "positive": positive_day,
                "negative": negative_day,
                "neutral": max(total_day - positive_day - negative_day, 0),
            }
        )

    positive = sum(item["positive"] for item in sentiment_trend)
    negative = sum(item["negative"] for item in sentiment_trend)

    return {
        "mode": "live",
        "brand": brand,
        "dateRange": {"from": dates[0], "to": dates[-1], "days": days},
        "kpis": {
            "mentions": count_data.get("total", sum(item["mentions"] for item in mentions_trend)),
            "reach": reach_data.get(
                "total",
                reach_data.get("social_media_reach_total", 0) + reach_data.get("non_social_media_reach_total", 0),
            ),
            "sentimentScore": round(((positive - negative) / max(positive + negative, 1)) * 100),
            "shareOfVoice": None,
            "engagementLift": None,
        },
        "mentionsTrend": mentions_trend,
        "sentimentTrend": sentiment_trend,
        "reachTrend": reach_trend,
        "sources": [],
        "topics": topics_data.get("topics", []),
        "hashtags": [],
        "links": [],
        "aiSummary": summary_data.get("summary", ""),
        "insights": [
            value
            for value in [
                insights_data.get("headline"),
                insights_data.get("trends"),
                insights_data.get("insights"),
                insights_data.get("recommendations"),
            ]
            if value
        ],
        "hotHours": [],
    }


@app.get("/health")
def health() -> Dict[str, str]:
    return {"status": "ok"}


@app.get("/api/auth/status")
def trial_auth_status(request: Request) -> Response:
    return JSONResponse(
        content={
            "required": bool(TRIAL_ACCESS_PASSWORD),
            "authenticated": _is_trial_authenticated(request),
        },
        headers={"Cache-Control": "no-store"},
    )


@app.post("/api/auth/login")
def trial_auth_login(payload: PasswordLogin, request: Request) -> Response:
    _check_rate_limit("auth_login", _client_id(request))
    if TRIAL_ACCESS_PASSWORD and not hmac.compare_digest(payload.password, TRIAL_ACCESS_PASSWORD):
        raise HTTPException(status_code=401, detail="The access password is incorrect.")

    response = JSONResponse(
        content={"authenticated": True},
        headers={"Cache-Control": "no-store"},
    )
    if TRIAL_ACCESS_PASSWORD:
        max_age = TRIAL_SESSION_HOURS * 60 * 60
        expires_at = int(time()) + max_age
        response.set_cookie(
            key=AUTH_COOKIE_NAME,
            value=_trial_session_value(expires_at),
            max_age=max_age,
            httponly=True,
            secure=AUTH_COOKIE_SECURE,
            samesite=AUTH_COOKIE_SAMESITE,
            path="/",
        )
    return response


@app.post("/api/auth/logout")
def trial_auth_logout() -> Response:
    response = JSONResponse(
        content={"authenticated": False},
        headers={"Cache-Control": "no-store"},
    )
    response.delete_cookie(
        key=AUTH_COOKIE_NAME,
        path="/",
        secure=AUTH_COOKIE_SECURE,
        samesite=AUTH_COOKIE_SAMESITE,
    )
    return response


@app.get("/api/projects")
def projects(
    request: Request,
    refresh: bool = Query(False),
) -> Dict[str, Any]:
    _check_rate_limit("projects_read", _client_id(request))
    project_list = _brand24_projects(force_refresh=refresh)
    return {
        "projects": project_list,
        "source": "monitoring-service",
        "syncedAt": int(time()),
    }


@app.get("/api/projects/{project_id}/briefing")
def project_briefing(
    project_id: str,
    request: Request,
    days: int = Query(7, ge=1, le=31),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
) -> Dict[str, Any]:
    _check_rate_limit("dashboard", _client_id(request))
    project = next((item for item in _brand24_projects() if item["id"] == project_id), None)
    if not project:
        raise HTTPException(status_code=404, detail="Project is not available in the connected monitoring account.")

    parsed_to = _parse_date_parameter(date_to, "date_to") or date.today()
    parsed_from = _parse_date_parameter(date_from, "date_from") or (parsed_to - timedelta(days=days - 1))
    if parsed_to < parsed_from:
        raise HTTPException(status_code=400, detail="date_to must be equal to or after date_from.")
    if (parsed_to - parsed_from).days > 30:
        raise HTTPException(status_code=400, detail="Briefing date ranges cannot exceed 31 days.")
    if parsed_to > date.today():
        raise HTTPException(status_code=400, detail="date_to cannot be in the future.")

    return _cached_dashboard(
        f"briefing:{project_id}:{parsed_from.isoformat()}:{parsed_to.isoformat()}",
        lambda: _live_briefing(project, parsed_from, parsed_to),
    )


@app.get("/api/reference/mention-categories")
def mention_categories(request: Request) -> Dict[str, Any]:
    _check_rate_limit("projects_read", _client_id(request))
    return {"categories": _brand24_mention_categories()}


@app.get("/api/projects/{project_id}/mentions")
def project_mentions(
    project_id: str,
    request: Request,
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    limit: int = Query(25, ge=1, le=100),
    cursor: Optional[str] = Query(None, max_length=2048),
    sentiment: Optional[str] = Query(None),
    category: Optional[str] = Query(None),
) -> Dict[str, Any]:
    _check_rate_limit("mentions_read", _client_id(request))
    project = next((item for item in _brand24_projects() if item["id"] == project_id), None)
    if not project:
        raise HTTPException(status_code=404, detail="Project is not available in the connected monitoring account.")

    parsed_to = _parse_date_parameter(date_to, "date_to") or date.today()
    parsed_from = _parse_date_parameter(date_from, "date_from") or (parsed_to - timedelta(days=6))
    if parsed_to < parsed_from:
        raise HTTPException(status_code=400, detail="date_to must be equal to or after date_from.")
    if (parsed_to - parsed_from).days > 30:
        raise HTTPException(status_code=400, detail="Mention explorer date ranges cannot exceed 31 days.")

    normalized_sentiment = sentiment.casefold().strip() if sentiment else None
    if normalized_sentiment and normalized_sentiment not in {"positive", "neutral", "negative"}:
        raise HTTPException(status_code=400, detail="sentiment must be positive, neutral, or negative.")

    normalized_category = category.casefold().strip() if category else None
    if normalized_category and normalized_category not in _brand24_mention_categories():
        raise HTTPException(status_code=400, detail="Category is not supported by the monitoring service.")

    return _live_mentions(
        project=project,
        date_from=parsed_from,
        date_to=parsed_to,
        limit=limit,
        cursor=cursor,
        sentiment=normalized_sentiment,
        category=normalized_category,
    )


@app.get("/api/projects/{project_id}/sources")
def project_sources(
    project_id: str,
    request: Request,
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
) -> Dict[str, Any]:
    _check_rate_limit("sources_read", _client_id(request))
    project = next((item for item in _brand24_projects() if item["id"] == project_id), None)
    if not project:
        raise HTTPException(status_code=404, detail="Project is not available in the connected monitoring account.")

    parsed_to = _parse_date_parameter(date_to, "date_to") or date.today()
    parsed_from = _parse_date_parameter(date_from, "date_from") or (parsed_to - timedelta(days=6))
    if parsed_to < parsed_from:
        raise HTTPException(status_code=400, detail="date_to must be equal to or after date_from.")
    if (parsed_to - parsed_from).days > 30:
        raise HTTPException(status_code=400, detail="Source intelligence date ranges cannot exceed 31 days.")

    return _cached_dashboard(
        f"sources:{project_id}:{parsed_from.isoformat()}:{parsed_to.isoformat()}",
        lambda: _live_sources(project, parsed_from, parsed_to),
    )


@app.get("/api/projects/{project_id}/topics")
def project_topics(
    project_id: str,
    request: Request,
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
) -> Dict[str, Any]:
    _check_rate_limit("topics_read", _client_id(request))
    project = next((item for item in _brand24_projects() if item["id"] == project_id), None)
    if not project:
        raise HTTPException(status_code=404, detail="Project is not available in the connected monitoring account.")

    parsed_to = _parse_date_parameter(date_to, "date_to") or date.today()
    parsed_from = _parse_date_parameter(date_from, "date_from") or (parsed_to - timedelta(days=6))
    if parsed_to < parsed_from:
        raise HTTPException(status_code=400, detail="date_to must be equal to or after date_from.")
    if (parsed_to - parsed_from).days > 30:
        raise HTTPException(status_code=400, detail="Topic intelligence date ranges cannot exceed 31 days.")

    return _cached_dashboard(
        f"topics:{project_id}:{parsed_from.isoformat()}:{parsed_to.isoformat()}",
        lambda: _live_topics(project, parsed_from, parsed_to),
    )


@app.post("/api/projects/{project_id}/exports")
def create_project_export(project_id: str, payload: ExportCreate, request: Request) -> Response:
    _check_rate_limit("exports_create", _client_id(request))
    project = next((item for item in _brand24_projects() if item["id"] == project_id), None)
    if not project:
        raise HTTPException(status_code=404, detail="Project is not available in the connected monitoring account.")
    normalized_format = payload.format.casefold()
    if normalized_format not in {"xlsx", "pptx", "pdf"}:
        raise HTTPException(status_code=400, detail="format must be xlsx, pptx, or pdf.")
    if payload.language.casefold() not in {"en", "ms"}:
        raise HTTPException(status_code=400, detail="language must be en or ms.")
    if payload.dateTo < payload.dateFrom:
        raise HTTPException(status_code=400, detail="dateTo must be equal to or after dateFrom.")
    if (payload.dateTo - payload.dateFrom).days > 30:
        raise HTTPException(status_code=400, detail="Export date ranges cannot exceed 31 days.")

    allowed_sections = {"overview", "daily", "mentions", "sources", "topics", "authors", "links", "hashtags"}
    normalized_sections = list(dict.fromkeys(section.casefold().strip() for section in payload.sections))
    unsupported_sections = sorted(set(normalized_sections) - allowed_sections)
    if unsupported_sections:
        raise HTTPException(status_code=400, detail=f"Unsupported export sections: {', '.join(unsupported_sections)}.")
    payload.sections = normalized_sections
    payload.language = payload.language.casefold()

    dataset = _build_export_dataset(project, payload.dateFrom, payload.dateTo, payload.sections)
    safe_project = re.sub(r"[^a-zA-Z0-9]+", "-", project["name"]).strip("-").lower() or "project"
    if normalized_format == "pptx":
        content = _build_export_presentation(dataset, payload)
        media_type = "application/vnd.openxmlformats-officedocument.presentationml.presentation"
    elif normalized_format == "pdf":
        content = build_pdf_report(
            dataset,
            {
                "language": payload.language,
                "reportTitle": payload.reportTitle,
                "organization": payload.organization,
                "sections": payload.sections,
            },
        )
        media_type = "application/pdf"
    else:
        content = _build_export_workbook(dataset, payload)
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    filename = f"{safe_project}-{payload.dateFrom.isoformat()}-{payload.dateTo.isoformat()}.{normalized_format}"
    return Response(
        content=content,
        media_type=media_type,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
